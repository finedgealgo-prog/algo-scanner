"""
Scanner Backtest Engine
-----------------------
Ported from sigma-backtest with scanner MongoDB collections:
  - scanner_stock_historical_data  (was: stock_historical_data)
  - scanner_index_historical_data  (was: index_historical_data)
  - scanner_stocks_list            (was: angel_stock_list)
"""

from __future__ import annotations

import calendar
import math
import os
import re
import time
import warnings
from collections import Counter
from datetime import datetime

import numpy as np
import numexpr as ne
import pandas as pd
from pymongo import MongoClient

warnings.simplefilter(action="ignore", category=pd.errors.SettingWithCopyWarning)

# ===============================
# CONFIG — same collections as sigma-backtest for identical results
# ===============================
MONGO_URI = "mongodb://localhost:27017"
DB_NAME = "stock_data"
COL_INDEX = "scanner_index_historical_data"
COL_STOCKS = "scanner_stock_historical_data"
COL_STOCK_LIST = "scanner_stock_list"

DEFAULT_UNIVERSES = ["nifty_50", "nifty_500"]
FORMULA = "(60% * 6 Month Performance + 30% * 3 Month Performance + 10% * 9 Month Performance ) / (3 Month Volatility  + 1 Month Volatility)"
FILTER_ACTION = "go_cash"
SAVE_SCORE_SNAPSHOTS = False
BUY_RENBALANCE_SCORE_SNAPSHOT_STATUS = False


# ===============================
# MongoDB — singleton connection
# ===============================
_mongo_client: MongoClient | None = None

def _get_client() -> MongoClient:
    global _mongo_client
    if _mongo_client is None:
        _mongo_client = MongoClient(MONGO_URI)
    return _mongo_client

def _mongo_col(collection: str):
    return _get_client()[DB_NAME][collection]


def load_collection_as_df(collection, query, projection, index_col=None):
    docs = list(_mongo_col(collection).find(query, projection))
    if not docs:
        return pd.DataFrame()
    df = pd.DataFrame(docs)
    if index_col and index_col in df.columns:
        df[index_col] = pd.to_datetime(df[index_col])
        df = df.sort_values(index_col).set_index(index_col)
    return df


def ensure_datetime_index(values):
    """Return a sorted DatetimeIndex so date comparisons don't hit ndarray/Timestamp issues."""
    dt_index = pd.DatetimeIndex(pd.to_datetime(values, errors="coerce"))
    if len(dt_index) == 0:
        return dt_index
    return dt_index[dt_index.notna()].sort_values().unique()


def get_latest_index_price(df_index, date, price_col="i_close_price"):
    sub = df_index[df_index.index <= date]
    if sub.empty:
        return None
    return sub[price_col].iloc[-1]


def load_index_daily(symbol, start, end):
    query = {
        "i_symbol": symbol,
        "ih_timestamp": {
            "$gte": pd.Timestamp(start).strftime("%Y-%m-%d"),
            "$lte": pd.Timestamp(end).strftime("%Y-%m-%d"),
        },
    }
    proj = {
        "_id": 0,
        "i_symbol": 1,
        "ih_timestamp": 1,
        "i_open_price": 1,
        "i_high_price": 1,
        "i_low_price": 1,
        "i_close_price": 1,
    }
    return load_collection_as_df(COL_INDEX, query, proj, index_col="ih_timestamp")


def load_stocks_bulk(symbols, start, end, min_price=None, max_price=None):
    def safe_float(val):
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    min_price = safe_float(min_price) or None
    max_price = safe_float(max_price) or None

    query = {
        "h_symbol": {"$in": symbols},
        "ch_timestamp": {
            "$gte": pd.Timestamp(start).strftime("%Y-%m-%d"),
            "$lte": pd.Timestamp(end).strftime("%Y-%m-%d"),
        },
    }
    proj = {
        "_id": 0,
        "h_symbol": 1,
        "ch_timestamp": 1,
        "ch_opening_price": 1,
        "ch_high_price": 1,
        "ch_low_price": 1,
        "ch_closing_price": 1,
    }

    df = load_collection_as_df(COL_STOCKS, query, proj)
    if df.empty:
        return df

    df["ch_timestamp"] = pd.to_datetime(df.get("ch_timestamp", df.index), errors="coerce")
    if "ch_timestamp" in df.columns:
        df = df.dropna(subset=["ch_timestamp"])
        df = df.sort_values(["h_symbol", "ch_timestamp"])
    df["ch_closing_price"] = pd.to_numeric(df["ch_closing_price"], errors="coerce").astype(float)

    df = (
        df.reset_index(drop=True)
        .groupby(["h_symbol", "ch_timestamp"], as_index=False)
        .agg(
            {
                "ch_opening_price": "first",
                "ch_high_price": "max",
                "ch_low_price": "min",
                "ch_closing_price": "last",
            }
        )
    )

    if min_price is not None or max_price is not None:
        latest = df.groupby("h_symbol", sort=False)["ch_closing_price"].last().reset_index()
        if min_price is not None:
            latest = latest[latest["ch_closing_price"] >= min_price]
        if max_price is not None:
            latest = latest[latest["ch_closing_price"] <= max_price]
        df = df[df["h_symbol"].isin(latest["h_symbol"].tolist())]

    df = df.sort_values(["h_symbol", "ch_timestamp"]).set_index("ch_timestamp")
    return df


def _parse_universe_date(date_str: str):
    """Convert 'DD-MM-YYYY' or 'YYYY-MM-DD' to pd.Timestamp. Returns None if empty."""
    if not date_str or str(date_str).strip() == "":
        return None
    s = str(date_str).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return pd.Timestamp(datetime.strptime(s, fmt))
        except ValueError:
            continue
    return None


def get_universe_stocks(universes=None):
    """Load all stocks from active scanner_universe_stock_list records."""
    if universes is None:
        universes = ["nifty_50"]
    col = _mongo_col("scanner_universe_stock_list")
    docs = list(col.find(
        {"filter_symbol": {"$in": universes}},
        {"_id": 0, "stocks": 1},
    ))
    symbols: set = set()
    for d in docs:
        for s in (d.get("stocks") or []):
            if s:
                symbols.add(s)
    return sorted(symbols)


# Cache universe records per filter_symbol to avoid repeated MongoDB queries
_universe_docs_cache: dict[str, list] = {}

def _load_universe_docs(universes: list[str]) -> list:
    key = ",".join(sorted(universes))
    if key not in _universe_docs_cache:
        col = _mongo_col("scanner_universe_stock_list")
        _universe_docs_cache[key] = list(col.find(
            {"filter_symbol": {"$in": universes}},
            {"_id": 0, "stocks": 1, "universe_period_start_date": 1, "universe_period_end_date": 1},
        ))
    return _universe_docs_cache[key]


def get_universe_stocks_at_date(universes, score_date: pd.Timestamp) -> list[str]:
    """Return stocks active in the universe at score_date (cached)."""
    docs = _load_universe_docs(list(universes))
    symbols: set[str] = set()
    for d in docs:
        start = _parse_universe_date(d.get("universe_period_start_date", ""))
        end_raw = d.get("universe_period_end_date", "")
        end = _parse_universe_date(end_raw)
        if start and start > score_date:
            continue
        if end and end < score_date:
            continue
        for s in (d.get("stocks") or []):
            if s:
                symbols.add(str(s).strip())
    return sorted(symbols)


# ===============================
# Precompute rolling metrics once
# ===============================
def precompute_rolling_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compute perf/vol metrics ONCE for the entire dataset.
    This avoids recomputing on every rebalance date — the biggest speedup.
    """
    df = df.sort_index().copy()
    df["ch_closing_price"] = pd.to_numeric(df["ch_closing_price"], errors="coerce")
    df["ret"] = df.groupby("h_symbol", group_keys=False)["ch_closing_price"].pct_change()

    periods = {"1M": 21, "3M": 63, "6M": 126, "9M": 189, "1Y": 252}
    for suffix, days in periods.items():
        df[f"perf_{suffix}"] = (
            df.groupby("h_symbol", group_keys=False)["ch_closing_price"]
              .pct_change(periods=days) * 100
        )
        df[f"vol_{suffix}"] = (
            df.groupby("h_symbol", group_keys=False)["ret"]
              .transform(lambda x: x.rolling(days, min_periods=5).std() * np.sqrt(252) * 100)
        )

    metric_cols = [c for c in df.columns if c.startswith(("perf_", "vol_"))]
    df[metric_cols] = df[metric_cols].fillna(0)
    df.drop(columns=["ret"], errors="ignore", inplace=True)
    return df


def fast_sigma_snapshot(ev_date: pd.Timestamp, df_precomputed: pd.DataFrame,
                         expr: str, active_symbols: list[str] | None = None) -> pd.DataFrame:
    """
    Fast score snapshot using precomputed metrics — no rolling recalculation.
    Filters to 400/401-day window matching sigma-backtest logic.
    """
    import calendar as _cal
    hist_days = 401 if _cal.isleap(ev_date.year - 1) else 400
    hist_start = ev_date - pd.Timedelta(days=hist_days)
    hist_end   = ev_date - pd.Timedelta(days=1)

    df_hist = df_precomputed.loc[
        (df_precomputed.index >= hist_start) & (df_precomputed.index <= hist_end)
    ]
    if df_hist.empty:
        return pd.DataFrame()

    if active_symbols:
        df_hist = df_hist[df_hist["h_symbol"].isin(active_symbols)]
        if df_hist.empty:
            return pd.DataFrame()

    # Get last record per symbol in the window
    idx = df_hist.groupby("h_symbol", sort=False).indices
    last_pos = [v[-1] for v in idx.values()]
    snap = df_hist.iloc[last_pos].copy()

    # Score using numexpr (fast)
    metric_cols = [c for c in snap.columns if c.startswith(("perf_", "vol_"))]
    local_dict = {c: snap[c].values / 100.0 for c in metric_cols}
    try:
        snap["score"] = ne.evaluate(expr, local_dict=local_dict)
    except Exception:
        snap["score"] = snap.eval(expr)
    snap["score"] = snap["score"].replace([np.inf, -np.inf], np.nan).fillna(0).round(6)

    snap = snap.set_index("h_symbol")
    snap["rank"] = snap["score"].rank(ascending=False, method="first").astype(int)
    snap = snap.sort_values("rank")

    expected = [
        "ch_opening_price", "ch_high_price", "ch_low_price", "ch_closing_price",
        "perf_1M", "perf_3M", "perf_6M", "perf_9M", "perf_1Y",
        "vol_1M", "vol_3M", "vol_6M", "vol_9M", "vol_1Y", "score", "rank",
    ]
    for c in expected:
        if c not in snap.columns:
            snap[c] = np.nan
    return snap[expected]


# ===============================
# Formula Parser
# ===============================
def parse_dynamic_formula(formula_str: str) -> str:
    expr = formula_str.replace(" ", " ")
    expr = re.sub(r"\s+", " ", expr).strip()
    if expr.count("(") != expr.count(")"):
        if expr.count("(") > expr.count(")"):
            expr += ")" * (expr.count("(") - expr.count(")"))
    expr = re.sub(r"(\d+(?:\.\d+)?)\s*%", lambda m: str(float(m.group(1)) / 100.0), expr)
    replacements = {
        "1 Month Performance": "perf_1M",
        "3 Month Performance": "perf_3M",
        "6 Month Performance": "perf_6M",
        "9 Month Performance": "perf_9M",
        "1 Year Performance": "perf_1Y",
        "1 Month Volatility": "vol_1M",
        "3 Month Volatility": "vol_3M",
        "6 Month Volatility": "vol_6M",
        "9 Month Volatility": "vol_9M",
        "1 Year Volatility": "vol_1Y",
    }
    for pat, col in sorted(replacements.items(), key=lambda x: -len(x[0])):
        expr = re.sub(pat, col, expr, flags=re.IGNORECASE)
    # Strip any stray identifier words that are NOT valid metric column names.
    # These arise when a stock name, strategy name, or unknown text leaks into
    # the formula string (e.g. "MAHABANK (0.4 * 1 Year Performance …)").
    # Leaving them causes numexpr/pandas to treat the word as a function call
    # and raise 'VariableNode' is not callable.
    _valid_ids = set(replacements.values())  # perf_1M … vol_1Y
    expr = re.sub(
        r"\b([A-Za-z][A-Za-z0-9_]*)\b",
        lambda m: m.group(0) if m.group(0) in _valid_ids else "",
        expr,
    )
    expr = re.sub(r"(?<=\d)\s+(?=[A-Za-z_(])", "*", expr)
    expr = re.sub(r"\s*([+\-/*()])\s*", r" \1 ", expr)
    expr = re.sub(r"\s+", " ", expr).strip()
    print(f"[Formula Parsed] {expr}")
    return expr


# ===============================
# Sigma Score Core
# ===============================
def compute_sigma_scores_core(df_stocks, ev_date, expr, df_meta=None, mode="backtest", score_model="current"):
    year = ev_date.year
    hist_days = 401 if calendar.isleap(year - 1) else 400
    hist_start = ev_date - pd.Timedelta(days=hist_days)
    hist_end = ev_date - pd.Timedelta(days=1)

    if not isinstance(df_stocks.index, pd.DatetimeIndex):
        if "ch_timestamp" in df_stocks.columns:
            df_stocks["ch_timestamp"] = pd.to_datetime(df_stocks["ch_timestamp"], errors="coerce")
            df_stocks = df_stocks.dropna(subset=["ch_timestamp"]).set_index("ch_timestamp").sort_index()
        else:
            raise ValueError("df_stocks must have DatetimeIndex or 'ch_timestamp' column")

    df_hist = df_stocks.loc[(df_stocks.index >= hist_start) & (df_stocks.index <= hist_end)].copy()
    if df_hist.empty:
        return pd.DataFrame()

    df_hist["ch_closing_price"] = pd.to_numeric(df_hist["ch_closing_price"], errors="coerce")
    df_hist["ret"] = df_hist.groupby("h_symbol", group_keys=False)["ch_closing_price"].pct_change()

    perf_periods = {"perf_1M": 21, "perf_3M": 63, "perf_6M": 126, "perf_9M": 189, "perf_1Y": 252}
    vol_periods = {"vol_1M": 21, "vol_3M": 63, "vol_6M": 126, "vol_9M": 189, "vol_1Y": 252}

    for name, days in perf_periods.items():
        df_hist[name] = df_hist.groupby("h_symbol")["ch_closing_price"].pct_change(periods=days) * 100
    _vol_factor = np.sqrt(252) if str(score_model or "current").strip().lower() != "old" else 1.0
    for name, days in vol_periods.items():
        df_hist[name] = df_hist.groupby("h_symbol")["ret"].transform(
            lambda x, d=days, f=_vol_factor: x.rolling(d, min_periods=5).std() * f * 100
        )
    for c in df_hist.columns:
        if c.startswith(("perf_", "vol_")):
            df_hist[c] = df_hist[c].fillna(0)

    local_dict = {col: df_hist[col].values / 100.0 for col in df_hist.columns if col.startswith(("perf_", "vol_"))}
    try:
        df_hist["score"] = ne.evaluate(expr, local_dict=local_dict)
    except Exception as e:
        print(f"[Fallback eval: {e}]")
        df_hist["score"] = df_hist.eval(expr)
    df_hist["score"] = df_hist["score"].replace([np.inf, -np.inf], np.nan).fillna(0).round(6)

    idx_dict = df_hist.groupby("h_symbol", sort=False).indices
    last_pos = [v[-1] for v in idx_dict.values()]
    snap = df_hist.iloc[last_pos].copy().drop(columns=["ret"], errors="ignore")

    if "h_symbol" not in snap.columns:
        snap["h_symbol"] = list(idx_dict.keys())

    if mode == "backtest":
        snap = snap.set_index("h_symbol")
        snap["rank"] = snap["score"].rank(ascending=False, method="first").astype(int)
        snap = snap.sort_values("rank")
        cols = [
            "ch_opening_price", "ch_high_price", "ch_low_price", "ch_closing_price",
            "perf_1M", "perf_3M", "perf_6M", "perf_9M", "perf_1Y",
            "vol_1M", "vol_3M", "vol_6M", "vol_9M", "vol_1Y", "score", "rank",
        ]
        for c in cols:
            if c not in snap.columns:
                snap[c] = np.nan
        return snap[cols]

    elif mode == "sigma":
        if df_meta is not None and not df_meta.empty:
            snap = snap.merge(df_meta[["symbol", "sector", "universe"]], left_on="h_symbol", right_on="symbol", how="left")
        else:
            snap["symbol"] = snap["h_symbol"]
            snap["sector"] = np.nan
            snap["universe"] = np.nan
        snap["last_price"] = snap["ch_closing_price"].round(2)
        snap["rank"] = snap["score"].rank(ascending=False, method="first").astype(int)
        snap = snap.sort_values("rank")
        cols = [
            "rank", "symbol", "sector", "universe", "score", "last_price",
            "perf_1M", "perf_3M", "perf_6M", "perf_9M", "perf_1Y",
            "vol_1M", "vol_3M", "vol_6M", "vol_9M", "vol_1Y",
        ]
        for c in cols:
            if c not in snap.columns:
                snap[c] = np.nan
        return snap[cols]
    else:
        raise ValueError("Invalid mode — use 'backtest' or 'sigma'")


# ===============================
# Regime Indicators
# ===============================
def compute_weekly_supertrend_fast(df_index_daily, atr_period=1, multiplier=2.5):
    wk = df_index_daily.resample("W-MON", label="left", closed="left").agg(
        {"i_open_price": "first", "i_high_price": "max", "i_low_price": "min", "i_close_price": "last"}
    ).dropna()
    hi, lo, cl = wk["i_high_price"].values, wk["i_low_price"].values, wk["i_close_price"].values
    n = len(wk)
    tr = np.maximum(hi - lo, np.maximum(np.abs(hi - np.roll(cl, 1)), np.abs(lo - np.roll(cl, 1))))
    tr[0] = hi[0] - lo[0]
    atr = pd.Series(tr).rolling(atr_period).mean().to_numpy()
    upper = ((hi + lo) / 2.0) + multiplier * atr
    lower = ((hi + lo) / 2.0) - multiplier * atr
    final_upper, final_lower = upper.copy(), lower.copy()
    for i in range(1, n):
        final_upper[i] = min(upper[i], final_upper[i - 1]) if cl[i - 1] <= final_upper[i - 1] else upper[i]
        final_lower[i] = max(lower[i], final_lower[i - 1]) if cl[i - 1] >= final_lower[i - 1] else lower[i]
    supertrend = np.full(n, np.nan)
    trend = np.ones(n, dtype=bool)
    for i in range(1, n):
        if cl[i] > final_upper[i - 1]:
            trend[i] = True
        elif cl[i] < final_lower[i - 1]:
            trend[i] = False
        else:
            trend[i] = trend[i - 1]
        supertrend[i] = final_lower[i] if trend[i] else final_upper[i]
    df_w = wk.copy()
    df_w["supertrend"] = supertrend
    df_w["signal"] = np.where(df_w["i_close_price"] > df_w["supertrend"], "Buy", "Sell")
    return df_w.dropna()


def compute_weekly_macd_fast(df_index_daily, fast=50, slow=100, signal=15):
    wk = df_index_daily.resample("W-MON", label="left", closed="left").agg(
        {"i_open_price": "first", "i_high_price": "max", "i_low_price": "min", "i_close_price": "last"}
    ).dropna()
    close = wk["i_close_price"]
    macd_line = close.ewm(span=fast, adjust=False).mean() - close.ewm(span=slow, adjust=False).mean()
    signal_line = macd_line.ewm(span=signal, adjust=False).mean()
    wk["macd_line"] = macd_line
    wk["macd_signal"] = signal_line
    wk["signal"] = np.where(wk["macd_line"] > wk["macd_signal"], "Buy", "Sell")
    return wk.dropna()


def compute_weekly_parabolic_sar_fast(df_index_daily, step=0.02, max_step=0.2):
    wk = df_index_daily.resample("W-MON", label="left", closed="left").agg(
        {"i_open_price": "first", "i_high_price": "max", "i_low_price": "min", "i_close_price": "last"}
    ).dropna()
    if wk.empty:
        return wk
    hi, lo, cl = wk["i_high_price"].to_numpy(), wk["i_low_price"].to_numpy(), wk["i_close_price"].to_numpy()
    n = len(wk)
    psar = np.empty(n); psar[:] = np.nan
    trend = np.zeros(n, dtype=bool)
    if n < 2:
        return wk.assign(psar=np.nan, signal="Sell")
    trend[0] = cl[1] > cl[0]
    ep = cl[1] if trend[0] else cl[1]
    psar[1] = lo[0] if trend[0] else hi[0]
    af = step
    prev_sar = float(psar[1])
    for i in range(1, n):
        prev_i = i - 1
        cur_sar = prev_sar if i == 1 else prev_sar + af * (ep - prev_sar)
        if trend[prev_i]:
            bound = lo[prev_i] if prev_i - 1 < 0 else min(lo[prev_i], lo[prev_i - 1])
            if cur_sar > bound:
                cur_sar = bound
        else:
            bound = hi[prev_i] if prev_i - 1 < 0 else max(hi[prev_i], hi[prev_i - 1])
            if cur_sar < bound:
                cur_sar = bound
        is_up = cl[i] > cur_sar
        if is_up and trend[prev_i]:
            if cl[i] > ep:
                ep = cl[i]; af = min(max_step, af + step)
            psar[i] = cur_sar; trend[i] = True
        elif (not is_up) and (not trend[prev_i]):
            if cl[i] < ep:
                ep = cl[i]; af = min(max_step, af + step)
            psar[i] = cur_sar; trend[i] = False
        else:
            psar[i] = ep; trend[i] = is_up; af = step; ep = cl[i]
        prev_sar = float(psar[i])
    if np.isnan(psar[0]):
        psar[0] = lo[0] if trend[1] else hi[0]
    wk = wk.copy()
    wk["psar"] = psar
    wk["signal"] = np.where(wk["i_close_price"] > wk["psar"], "Buy", "Sell")
    return wk.dropna()


# ===============================
# Rebalance Date Generator
# ===============================
def generate_rebalance_dates(df_index, start_date, end_date, rebalance_frequency, rebalance_day_or_week):
    start_date = pd.Timestamp(start_date)
    end_date = pd.Timestamp(end_date)
    freq = rebalance_frequency.strip().lower()
    dates = []
    weekday_map = {1: "MON", 2: "TUE", 3: "WED", 4: "THU", 5: "FRI"}
    weekday_code = weekday_map.get(int(rebalance_day_or_week), "MON")

    if freq == "monthly":
        for m in pd.date_range(start_date, end_date, freq="MS"):
            target = m + pd.DateOffset(days=int(rebalance_day_or_week) - 1)
            cand = df_index[df_index.index >= target].index.min()
            if pd.notna(cand) and cand <= end_date:
                dates.append(cand)
    elif freq == "weekly":
        for d in pd.date_range(start_date, end_date, freq=f"W-{weekday_code}"):
            cand = df_index[df_index.index >= d].index.min()
            if pd.notna(cand) and cand <= end_date:
                dates.append(cand)
    elif freq == "bi-weekly":
        all_days = pd.date_range(start_date, end_date, freq=f"W-{weekday_code}")[::2]
        for d in all_days:
            cand = df_index[df_index.index >= d].index.min()
            if pd.notna(cand) and cand <= end_date:
                dates.append(cand)
    elif freq == "bi-monthly":
        for m in pd.date_range(start_date, end_date, freq="MS"):
            target = m + pd.Timedelta(days=7) + pd.DateOffset(days=(rebalance_day_or_week - 1))
            cand = df_index[df_index.index >= target].index.min()
            if pd.notna(cand) and cand <= end_date:
                dates.append(cand)
    elif freq == "quarterly":
        for q in pd.date_range(start_date, end_date, freq="QS"):
            target = q + pd.DateOffset(days=int(rebalance_day_or_week) - 1)
            cand = df_index[df_index.index >= target].index.min()
            if pd.notna(cand) and cand <= end_date:
                dates.append(cand)
    else:
        return generate_rebalance_dates(df_index, start_date, end_date, "Monthly", rebalance_day_or_week)

    return sorted(set(dates))


# ===============================
# MAIN BACKTEST FUNCTION
# ===============================
def ultra_backtest_v2_fastest_fixed_v3(
    df_index,
    df_stocks,
    df_gold,
    start_date,
    end_date,
    top_n,
    initial_capital=1_000_000,
    gold_alloc=0.85,
    FORMULA="((70% * 6 Month Volatility) + (20% * 3 Month Performance) + (10% * 1 Year Performance)) / 3 Month Volatility",
    EXIT_RANK=18,
    EXIT_REGIME_FILTER_STATUS=True,
    UNCORRELATED_ASSET_STATUS=True,
    INDICATOR_NAME="supertrend",
    INDICATOR_PARAMS=(1, 2.5),
    FILTER_ACTION="go_cash",
    REBALANCE_FREQUENCY="Monthly",
    REBALANCE_DAY_OR_WEEK=15,
    EXECUTION_TIMING="same_day",  # "same_day" or "next_trading_day"
    REGIME_MONDAY_ONLY=False,              # True = check only on Monday | False = if Monday holiday, use next trading day same week
    UNIVERSE_INDEXES=None,
    SAVE_EXCEL=False,
    score_model="current",
):
    print("REBALANCE_FREQUENCY",REBALANCE_FREQUENCY,"INDICATOR_NAME",INDICATOR_NAME,"INDICATOR_PARAMS",INDICATOR_PARAMS)
    """
    Rules implemented:
    - Supertrend = Sell: if holding stocks → sell stocks and buy gold_bees (same/next trading day)
      and skip all rebalances while holding gold_bees.
    - Supertrend = Buy: if holding gold_bees → sell gold_bees then immediately buy stocks by score (same/next day).
    - Rebalance when holding stocks:
        Phase 1: SELL_REBALANCE, Phase 2: BUY_REBALANCE (same/next).
    Keeps existing outputs (cash_after, cumulative_value, remaining_capital, snapshots, metrics, etc.).
    """

    import numpy as np, pandas as pd, re, time, os
    from openpyxl import Workbook
    t0 = time.time()
    
    # ------------------------------------------------------
    # Expand stock history window by 500 days for score lookbacks
    # ------------------------------------------------------
    buffer_days = 500
    extended_start_date = pd.to_datetime(start_date) - pd.Timedelta(days=buffer_days)

    # Keep the stock data only from (extended_start_date → end_date)
    df_stocks = df_stocks.loc[(df_stocks.index >= extended_start_date) & (df_stocks.index <= end_date)].copy()


    # ------------------------------------------------------
    # Helper functions
    # ------------------------------------------------------
    def next_trading_day(d: pd.Timestamp) -> pd.Timestamp:
        ix = df_index.index
        pos = ix.searchsorted(d, side="right")
        return ix[pos] if pos < len(ix) else d

    def trade_day(d: pd.Timestamp, for_buy: bool) -> pd.Timestamp:
        """Return execution day depending on EXECUTION_TIMING."""
        if not for_buy:
            return d
        return d if EXECUTION_TIMING.lower() == "same_day" else next_trading_day(d)

    def portfolio_value_on(d):
        total = cash
        for s, q in holdings.items():
            px = get_latest_index_price(df_gold, d) if s == "gold_bees" else fast_price(s, d)
            total += (px or 0.0) * q
        return total

    def add_trade_row(d, symbol, action, qty, px,  rank=None):
        nonlocal cash
        if action.startswith("BUY"):
            cash -= qty * px
        elif action.startswith("SELL"):
            cash += qty * px
        cum = portfolio_value_on(d) #- {rank}
        ranks = str(rank) if rank is not None else "None"
        trades.append({
            "date": d, "symbol": f"{symbol}", "action": action,
            "shares": int(qty), "price": float(round(px, 6)),
            "cash_after": float(round(cash, 2)),
            "cumulative_value": float(round(cum, 2)),
            "remining_capital": float(round(cash, 2)),
            # "rank": int(rank) if rank is not None else None
        })
    
    # def add_trade_row(d, symbol, action, qty, px, rank=None):
    #     nonlocal cash

    #     # Update cash
    #     if action.startswith("BUY"):
    #         cash -= qty * px
    #     elif action.startswith("SELL"):
    #         cash += qty * px

    #     # Portfolio value
    #     cum = portfolio_value_on(d)

    #     # ⚡ FIX: Convert rank safely
    #     safe_rank = None if rank is None else int(rank)

    #     trades.append({
    #         "date": d,
    #         "symbol": symbol,        # 👈 don't mix rank here
    #         "rank": safe_rank,       # 👈 proper rank added
    #         "action": action,
    #         "shares": int(qty),
    #         "price": float(round(px, 6)),
    #         "cash_after": float(round(cash, 2)),
    #         "cumulative_value": float(round(cum, 2)),
    #         "remaining_capital": float(round(cash, 2)),
    #     })


    def push_snapshot(d):
        pos = []
        for s, q in sorted(holdings.items()):
            px = get_latest_index_price(df_gold, d) if s == "gold_bees" else fast_price(s, d)
            pos.append({"symbol": s, "qty": int(q), "price": float(px or 0.0), "value": float((px or 0.0)*q)})
        snapshots.append({
            "date": d,
            "cash": float(round(cash, 2)),
            "holdings_count": int(len(holdings)),
            "total_value": float(round(portfolio_value_on(d), 2)),
            "positions": pos
        })

    # ------------------------------------------------------
    # Parse scoring formula
    # ------------------------------------------------------
    expr = parse_dynamic_formula(FORMULA)

    # ------------------------------------------------------
    # Compute Supertrend regime signals
    # ------------------------------------------------------
    # ------------------------------------------------------
    # Compute regime signals (Supertrend or Parabolic SAR)
    # ------------------------------------------------------
    if EXIT_REGIME_FILTER_STATUS:
        name = INDICATOR_NAME.lower()
        if name == "supertrend":
            atr = INDICATOR_PARAMS[0] if not isinstance(INDICATOR_PARAMS, dict) else INDICATOR_PARAMS.get("atr_period", 1)
            mult = INDICATOR_PARAMS[1] if not isinstance(INDICATOR_PARAMS, dict) else INDICATOR_PARAMS.get("multiplier", 2.5)
            df_week = compute_weekly_supertrend_fast(df_index, atr, mult)
            sig_map = dict(zip(df_week.index, df_week["signal"]))
        elif name in ("parabolic_sar", "psar"):
            # INDICATOR_PARAMS can be tuple/list (step, max_step) or dict {"step":..., "max_step":...}
            if isinstance(INDICATOR_PARAMS, dict):
                step = INDICATOR_PARAMS.get("step", 0.02)
                max_step = INDICATOR_PARAMS.get("max_step", 0.2)
            else:
                try:
                    step, max_step = float(INDICATOR_PARAMS[0]), float(INDICATOR_PARAMS[1])
                except Exception:
                    step, max_step = 0.02, 0.2
            df_week = compute_weekly_parabolic_sar_fast(df_index, step=step, max_step=max_step)
            sig_map = dict(zip(df_week.index, df_week["signal"]))
        elif name == "macd":

            if isinstance(INDICATOR_PARAMS, dict):
                fast = INDICATOR_PARAMS.get("fast", 50)
                slow = INDICATOR_PARAMS.get("slow", 100)
                signal = INDICATOR_PARAMS.get("signal", 15)
            else:
                try:
                    fast, slow, signal = INDICATOR_PARAMS
                except:
                    fast, slow, signal = 50, 100, 15

            df_week = compute_weekly_macd_fast(df_index, fast, slow, signal)
            sig_map = dict(zip(df_week.index, df_week["signal"]))

        else:
            raise ValueError(f"Unsupported indicator: {INDICATOR_NAME}")

        # If REGIME_MONDAY_ONLY=False, remap any Monday holiday to next trading day in same week
        if not REGIME_MONDAY_ONLY and sig_map:
            trading_days = set(df_index.index)
            remapped: dict = {}
            for d, signal in sig_map.items():
                if d in trading_days:
                    remapped[d] = signal
                else:
                    # Find next trading day within same week (up to 6 days ahead)
                    shifted = df_index.index[df_index.index > d]
                    shifted = shifted[shifted <= d + pd.Timedelta(days=6)]
                    if len(shifted) > 0:
                        remapped[shifted[0]] = signal
                    # if no trading day found in that week, skip this week entirely
            sig_map = remapped
    else:
        sig_map = {}


    # ------------------------------------------------------
    # Generate rebalance dates
    # ------------------------------------------------------
    rebalance_dates_raw = generate_rebalance_dates(
        df_index, start_date, end_date, REBALANCE_FREQUENCY, REBALANCE_DAY_OR_WEEK
    )
    rebalance_dates_raw = sorted(set(rebalance_dates_raw))

    # ------------------------------------------------------
    # Prepare stock cache
    # ------------------------------------------------------
    df_stocks = df_stocks.sort_index()
    stock_cache = {s: g for s, g in df_stocks.groupby("h_symbol")}
    all_stock_dates = np.unique(df_stocks.index.values)

    # Pre-convert OHLC arrays for fast numpy slicing inside sigma_score_snapshot
    _price_cache: dict = {}
    for _sym, _sdf in stock_cache.items():
        _idx = _sdf.index
        _close = pd.to_numeric(_sdf["ch_closing_price"], errors="coerce").values.astype(float)
        _open  = pd.to_numeric(_sdf["ch_opening_price"], errors="coerce").values.astype(float)
        _high  = pd.to_numeric(_sdf["ch_high_price"],    errors="coerce").values.astype(float)
        _low   = pd.to_numeric(_sdf["ch_low_price"],     errors="coerce").values.astype(float)
        _price_cache[_sym] = (_idx, _close, _open, _high, _low)

    def fast_price(sym, date):
        df = stock_cache.get(sym)
        if df is None:
            return None
        pos = df.index.searchsorted(date, side="right") - 1
        if pos < 0:
            return None
        # print("df.iloc[pos]",df.iloc[pos])
        return float(df.iloc[pos]["ch_closing_price"])

    # ------------------------------------------------------
    # Map effective rebalance dates
    # ------------------------------------------------------
    effective_rebalance_dates = []
    for raw in rebalance_dates_raw:
        nd = df_index[df_index.index >= raw].index.min()
        if pd.isna(nd):
            continue
        if nd not in all_stock_dates:
            pos = np.searchsorted(all_stock_dates, np.datetime64(nd), side="left")
            if pos >= len(all_stock_dates):
                continue
            nd = pd.Timestamp(all_stock_dates[pos])
        effective_rebalance_dates.append(nd)
    effective_rebalance_dates = sorted(set(effective_rebalance_dates))
    
    def clean_dates(dates):
        """Clean None/NaT and sort Timestamps."""
        return sorted(pd.to_datetime([d for d in dates if d is not None and pd.notna(d)]).unique())

    # Weekly supertrend (for regime entry/exit)
    regime_dates = clean_dates(sig_map.keys())

    # Monthly or configured rebalance dates
    rebalance_dates = clean_dates(effective_rebalance_dates)

    # Daily dates from stock data (for stoploss)
    if isinstance(df_stocks.index, pd.MultiIndex):
        daily_dates = clean_dates(df_stocks.index.get_level_values(0).unique())
    else:
        daily_dates = clean_dates(df_stocks.index.unique())

    # Limit daily range to backtest window
    daily_dates = [d for d in daily_dates if start_date <= d <= end_date]

    print(f"✅ Dates loaded → Regime={len(regime_dates)}, Rebalance={len(rebalance_dates)}, Daily={len(daily_dates)}")
    
    
    event_dates = (
        sorted(set(list(sig_map.keys()) + list(effective_rebalance_dates)))
        if sig_map
        else effective_rebalance_dates[:]
    )

    # ------------------------------------------------------
    # Setup Excel file for saving rebalance score snapshots
    # ------------------------------------------------------
    score_excel_path = "stock_scores_backtest.xlsx"
    if SAVE_SCORE_SNAPSHOTS:
        if not os.path.exists(score_excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws["A1"] = "Sigma Scoring Backtest Sheets"
            ws["A2"] = f"Created {pd.Timestamp.now()}"
            wb.save(score_excel_path)

    def save_score_snapshot_to_excel(date, snap_df):
        """Append rebalance snapshot to Excel as new sheet."""
        if not SAVE_SCORE_SNAPSHOTS:
            return  # ✅ Skip Excel writes entirely for performance
        try:
            if snap_df is None or snap_df.empty:
                return
            df_out = snap_df.reset_index().rename(columns={"index": "h_symbol"})
            if SAVE_SCORE_SNAPSHOTS:
                with pd.ExcelWriter(score_excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    sheet_name = str(pd.to_datetime(date).strftime("%Y-%m-%d"))
                    df_out.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        except Exception as e:
            print(f"[Warning] Could not save score sheet for {date}: {e}")

    # ------------------------------------------------------
    # Sigma scoring snapshot — fast numpy per-symbol, correct slice semantics
    # (matches sigma-backtest compute_sigma_scores_core exactly)
    # ------------------------------------------------------
    _snap_cache: dict = {}

    def sigma_score_snapshot(ev_date: pd.Timestamp):
        if ev_date in _snap_cache:
            return _snap_cache[ev_date]

        import calendar as _cal
        year = ev_date.year
        hist_days = 401 if _cal.isleap(year - 1) else 400
        hist_start = ev_date - pd.Timedelta(days=hist_days)
        hist_end   = ev_date - pd.Timedelta(days=1)

        hs_np = np.datetime64(hist_start)
        he_np = np.datetime64(hist_end)

        rows = []
        for sym, (idx, prices, opens, highs, lows) in _price_cache.items():
            sl = np.searchsorted(idx, hs_np, side="left")
            sr = np.searchsorted(idx, he_np, side="right")
            if sr - sl < 5:
                continue
            px   = prices[sl:sr]
            n    = len(px)

            # daily returns (length n-1)
            with np.errstate(divide="ignore", invalid="ignore"):
                rets = np.where(px[:-1] != 0, px[1:] / px[:-1] - 1, 0.0)

            def pc(d):
                return (px[-1] / px[-1 - d] - 1) * 100 if (n > d and px[-1 - d] != 0) else 0.0

            _bt_vol_factor = np.sqrt(252) if str(score_model or "current").strip().lower() != "old" else 1.0
            def rstd(d, _f=_bt_vol_factor):
                w = rets[max(0, len(rets) - d):]
                return float(np.std(w, ddof=1)) * _f * 100 if len(w) >= 5 else 0.0

            rows.append({
                "h_symbol":         sym,
                "ch_opening_price": float(opens[sr - 1]),
                "ch_high_price":    float(highs[sr - 1]),
                "ch_low_price":     float(lows[sr - 1]),
                "ch_closing_price": float(px[-1]),
                "perf_1M": pc(21),  "perf_3M": pc(63),  "perf_6M": pc(126),
                "perf_9M": pc(189), "perf_1Y": pc(252),
                "vol_1M":  rstd(21), "vol_3M":  rstd(63),  "vol_6M":  rstd(126),
                "vol_9M":  rstd(189),"vol_1Y":  rstd(252),
            })

        if not rows:
            return pd.DataFrame()

        snap = pd.DataFrame(rows).set_index("h_symbol")
        _mc  = ["perf_1M","perf_3M","perf_6M","perf_9M","perf_1Y",
                "vol_1M", "vol_3M", "vol_6M", "vol_9M", "vol_1Y"]
        local_dict = {c: snap[c].values / 100.0 for c in _mc}
        try:
            snap["score"] = ne.evaluate(expr, local_dict=local_dict)
        except Exception:
            snap["score"] = snap.eval(expr)
        snap["score"] = snap["score"].replace([np.inf, -np.inf], np.nan).fillna(0).round(6)
        snap["rank"]  = snap["score"].rank(ascending=False, method="first").astype(int)
        snap = snap.sort_values("rank")

        cols = ["ch_opening_price","ch_high_price","ch_low_price","ch_closing_price",
                "perf_1M","perf_3M","perf_6M","perf_9M","perf_1Y",
                "vol_1M","vol_3M","vol_6M","vol_9M","vol_1Y","score","rank"]
        for c in cols:
            if c not in snap.columns:
                snap[c] = np.nan
        save_score_snapshot_to_excel(ev_date, snap)
        result = snap[cols]
        _snap_cache[ev_date] = result
        return result

    # ===================================================
    # 4️⃣ Combine and Deduplicate Event Dates
    # ===================================================
    event_dates_raw = list(sig_map.keys()) + list(effective_rebalance_dates)
    event_dates = sorted(set(event_dates_raw))

    # Detect and log duplicates (for debug)
    from collections import Counter
    dupes = [d for d, c in Counter(event_dates_raw).items() if c > 1]
    if dupes:
        print(f"⚠️ Found duplicate event dates: {[d.strftime('%Y-%m-%d') for d in dupes]}")

    # Cache to skip duplicate compute_sigma_scores_core() calls
    computed_dates_cache = set()


    # ===================================================
    # 4️⃣ Combine and Deduplicate Event Dates (Fully Normalized)
    # ===================================================
    from collections import Counter

    event_dates_raw = list(sig_map.keys()) + list(effective_rebalance_dates)

    def normalize_date(ts):
        if pd.isna(ts):
            return None
        if not isinstance(ts, pd.Timestamp):
            ts = pd.Timestamp(ts)
        return ts.normalize()

    event_dates_normalized = [normalize_date(d) for d in event_dates_raw if d is not None]
    event_dates = sorted(set(event_dates_normalized))

    counts = Counter(event_dates_normalized)
    dupes = [d for d, c in counts.items() if c > 1]
    if dupes:
        print(f"⚠️ Found {len(dupes)} duplicate event dates after normalization: {[d.strftime('%Y-%m-%d') for d in dupes]}")

    computed_dates_cache = set()

    # ===================================================
    # 5️⃣ Init Portfolio
    # ===================================================
    cash = float(initial_capital)
    holdings = {}
    trades, equity, snapshots = [], [], []
    first_trading_day = df_index.index.min()

    if EXIT_REGIME_FILTER_STATUS:
        past_signals = [d for d in sorted(sig_map.keys()) if d <= start_date]
        initial_signal = sig_map[past_signals[-1]] if past_signals else "Buy"
    else:
        initial_signal = "Buy"

    state = initial_signal
    last_signal = initial_signal
    regime_events: list[dict] = []

    # ------------------------------------------------------
    # Main event loop
    # ------------------------------------------------------
    for ev_date in daily_dates: #event_dates #daily_dates
        # print(f"Processing event date: {ev_date.date()} | State: {state} | Cash: {cash:.2f} | Holdings: {len(holdings)}")
        # print("holdings",holdings)
        # Skip any pre-start scoring days (used only for Sigma lookback warm-up)
        if ev_date < pd.to_datetime(start_date):
            continue
        snap = None  # ✅ Always initialize at the top of loop to prevent UnboundLocalError

        # ==================================================
        # Regime Change: Supertrend Signal
        # ==================================================
        if EXIT_REGIME_FILTER_STATUS and ev_date in sig_map:
            print(f"[REGIME CHECK] Date: {ev_date.date()} | Signal: {sig_map[ev_date]} | Last Signal: {last_signal} | Changed: {sig_map[ev_date] != last_signal}")
        if EXIT_REGIME_FILTER_STATUS and ev_date in sig_map and sig_map[ev_date] != last_signal:
            sig = sig_map[ev_date]

            # ---------- SELL signal ----------
            if sig == "Sell":
                sell_d = trade_day(ev_date, False)
                print(f"  [SELL SIGNAL] Selling all stocks on {sell_d} | Holdings before: {list(holdings.keys())}")
                event_sold_stocks = []
                event_skipped_stocks = []
                for sym, qty in list(holdings.items()):
                    if sym == "gold_bees":
                        continue
                    px = fast_price(sym, sell_d)
                    if not px:
                        print(f"    [SKIP] {sym} — no price found, stays in holdings")
                        event_skipped_stocks.append({"symbol": sym, "reason": "no_price"})
                        continue
                    if FILTER_ACTION == "half_portfolio":
                        sell_qty = qty // 2
                        if sell_qty > 0:
                            add_trade_row(sell_d, sym, "SELL_HALF_SIGNAL", sell_qty, px)
                            remain = qty - sell_qty
                            if remain > 0:
                                holdings[sym] = remain
                            else:
                                holdings.pop(sym, None)
                            print(f"    [SELL HALF] {sym} qty={sell_qty} @ {px}")
                            event_sold_stocks.append({"symbol": sym, "qty": sell_qty, "price": float(px), "action": "SELL_HALF_SIGNAL"})
                    else:
                        rank = None
                        if snap is not None and sym in snap.index:
                            rank = int(snap.loc[sym, "rank"])
                        add_trade_row(sell_d, sym, "SELL_SIGNAL", qty, px, rank)
                        holdings.pop(sym, None)
                        print(f"    [SOLD] {sym} qty={qty} @ {px} | rank={rank}")
                        event_sold_stocks.append({"symbol": sym, "qty": qty, "price": float(px), "rank": rank, "action": "SELL_SIGNAL"})

                event_bought_gold = None
                # BUY GoldBees after selling
                if UNCORRELATED_ASSET_STATUS:
                    buy_d = trade_day(ev_date, True)
                    gpx = get_latest_index_price(df_gold, buy_d)
                    if gpx and gpx > 0:
                        qty = int((cash * gold_alloc) // gpx)
                        if qty > 0:
                            holdings["gold_bees"] = qty
                            add_trade_row(buy_d, "gold_bees", "BUY_GOLDBEES", qty, float(gpx))
                            print(f"  [BUY GOLDBEES] date={buy_d} qty={qty} @ {gpx} | cash_used={qty*gpx:.2f} | cash_remaining={cash:.2f}")
                            event_bought_gold = {"symbol": "gold_bees", "qty": qty, "price": float(gpx), "date": str(buy_d.date() if hasattr(buy_d, "date") else buy_d)}
                    else:
                        print(f"  [SKIP GOLDBEES] No gold price found on {buy_d}")
                else:
                    print(f"  [SKIP GOLDBEES] UNCORRELATED_ASSET_STATUS is OFF")

                regime_events.append({
                    "signal_date": str(ev_date.date() if hasattr(ev_date, "date") else ev_date),
                    "signal_day": ev_date.strftime("%A") if hasattr(ev_date, "strftime") else "",
                    "signal": "Sell",
                    "sell_date": str(sell_d.date() if hasattr(sell_d, "date") else sell_d),
                    "stocks_sold": event_sold_stocks,
                    "stocks_skipped": event_skipped_stocks,
                    "gold_bought": event_bought_gold,
                    "stocks_bought": [],
                    "gold_sold": None,
                })
                print(f"  [SELL SIGNAL DONE] Holdings after: {list(holdings.keys())} | Cash: {cash:.2f}")
                state = "Sell"
                last_signal = sig

            # ---------- BUY signal ----------
            elif sig == "Buy":
                print(f"  [BUY SIGNAL] Holdings before: {list(holdings.keys())} | Cash: {cash:.2f}")
                event_gold_sold = None
                # Sell GoldBees first
                if "gold_bees" in holdings and UNCORRELATED_ASSET_STATUS:
                    sell_d = trade_day(ev_date, False)
                    gpx = get_latest_index_price(df_gold, sell_d)
                    if gpx and gpx > 0:
                        qty = holdings.pop("gold_bees")
                        add_trade_row(sell_d, "gold_bees", "SELL_GOLDBEES", qty, float(gpx))
                        print(f"  [SOLD GOLDBEES] date={sell_d} qty={qty} @ {gpx} | cash after={cash:.2f}")
                        event_gold_sold = {"symbol": "gold_bees", "qty": qty, "price": float(gpx), "date": str(sell_d.date() if hasattr(sell_d, "date") else sell_d)}
                    else:
                        print(f"  [SKIP SELL GOLDBEES] No gold price on {sell_d}")
                else:
                    print(f"  [NO GOLDBEES] Nothing to sell (gold_bees not in holdings or UNCORRELATED_ASSET_STATUS OFF)")

                # Buy top stocks based on Sigma scoring
                buy_d = trade_day(ev_date, True)
                snap = sigma_score_snapshot(ev_date)
                event_bought_stocks = []
                if not snap.empty and top_n > 0:
                    top_syms = snap.head(top_n).index.tolist()
                    candidates = [s for s in top_syms if holdings.get(s, 0) == 0]
                    print(f"  [BUY REENTER] top_n={top_n} | top_syms={top_syms} | already_held={[s for s in top_syms if holdings.get(s,0)>0]} | candidates={candidates}")
                    if candidates and cash > 0:
                        alloc = cash / len(candidates)
                        buy_block = []
                        for sym in candidates:
                            p = fast_price(sym, buy_d)
                            if p and p > 0:
                                qty = int(alloc // p)
                                if qty > 0:
                                    holdings[sym] = qty
                                    buy_block.append((sym, qty, float(p)))
                                else:
                                    print(f"    [SKIP] {sym} — qty=0 (alloc={alloc:.2f}, price={p})")
                            else:
                                print(f"    [SKIP] {sym} — no price on {buy_d}")
                        for sym, qty, px in buy_block:
                            rank = int(snap.loc[sym, "rank"]) if "rank" in snap.columns else None
                            add_trade_row(buy_d, sym, "BUY_REENTER", qty, px, rank)
                            print(f"    [BOUGHT] {sym} qty={qty} @ {px} | rank={rank}")
                            event_bought_stocks.append({"symbol": sym, "qty": qty, "price": float(px), "rank": rank, "action": "BUY_REENTER"})
                    else:
                        print(f"  [SKIP BUY] candidates empty or cash=0")
                else:
                    print(f"  [SKIP BUY] snap empty or top_n=0")

                regime_events.append({
                    "signal_date": str(ev_date.date() if hasattr(ev_date, "date") else ev_date),
                    "signal_day": ev_date.strftime("%A") if hasattr(ev_date, "strftime") else "",
                    "signal": "Buy",
                    "buy_date": str(buy_d.date() if hasattr(buy_d, "date") else buy_d),
                    "gold_sold": event_gold_sold,
                    "stocks_bought": event_bought_stocks,
                    "stocks_sold": [],
                    "stocks_skipped": [],
                    "gold_bought": None,
                })
                print(f"  [BUY SIGNAL DONE] Holdings after: {list(holdings.keys())} | Cash: {cash:.2f}")
                state = "Buy"
                last_signal = sig

        # ==================================================
        # Rebalance Events (Only when holding stocks)
        # ==================================================
        if (ev_date in effective_rebalance_dates) and state == "Buy" and ("gold_bees" not in holdings):

            # compute fresh Sigma scores (400-day lookback)
            snap = sigma_score_snapshot(ev_date)
            if snap is None or snap.empty:
                continue

            # select top N by score
            top_syms = snap.head(top_n).index.tolist()

            # -------- Phase 1: SELL_REBALANCE — exit stocks below EXIT_RANK --------
            sell_d = trade_day(ev_date, False)
            sell_block = []
            for sym, qty in list(holdings.items()):
                if sym == "gold_bees":
                    continue
                if (sym not in snap.index) or (snap.loc[sym, "rank"] > EXIT_RANK):
                    px = fast_price(sym, sell_d)
                    holdings.pop(sym, None)  # always remove, even if no price
                    if px:
                        sell_block.append((sym, qty, px))

            for sym, qty, px in sell_block:
                rank = int(snap.loc[sym, "rank"]) if sym in snap.index and "rank" in snap.columns else None
                add_trade_row(sell_d, sym, "SELL_REBALANCE", qty, px, rank)

            # -------- Phase 2: BUY_REBALANCE — fill empty slots with top-N candidates --------
            buy_d = trade_day(ev_date, True)
            current_count = len(holdings)
            if BUY_RENBALANCE_SCORE_SNAPSHOT_STATUS:
                snap = sigma_score_snapshot(buy_d)
                if snap is None or snap.empty:
                    continue
                top_syms = snap.head(top_n).index.tolist()

            remaining_slots = top_n - current_count
            if remaining_slots > 0:
                candidates = [s for s in top_syms if s not in holdings][:remaining_slots]
                if candidates and cash > 0:
                    alloc = cash / len(candidates)
                    buy_block = []
                    for sym in candidates:
                        p = fast_price(sym, buy_d)
                        if p and p > 0:
                            qty = int(alloc // p)
                            if qty > 0:
                                holdings[sym] = qty
                                buy_block.append((sym, qty, float(p)))
                    for sym, qty, px in buy_block:
                        rank = int(snap.loc[sym, "rank"]) if "rank" in snap.columns else None
                        add_trade_row(buy_d, sym, "BUY_REBALANCE", qty, px, rank)

        # ==================================================
        # End of day snapshot & equity tracking
        # ==================================================
        equity.append({
            "date": ev_date,
            "value": float(round(portfolio_value_on(ev_date), 2)),
            "cash": float(round(cash, 2)),
            "holdings_count": int(len(holdings)),
            "state": state
        })
        push_snapshot(ev_date)
    # ------------------------------------------------------
    # Wrap up - Metrics, trade reconciliation, and outputs
    # ------------------------------------------------------
    df_eq = pd.DataFrame(equity).sort_values("date").reset_index(drop=True)
    df_tr = pd.DataFrame(trades)
    
    if not df_tr.empty:
        df_tr["date"] = pd.to_datetime(df_tr["date"])
        eod = df_tr.groupby("date")["cash_after"].last().rename("EOD").reset_index()
        df_tr = df_tr.merge(eod, on="date", how="left")
        df_tr["remining_capital"] = df_tr["EOD"].round(2)
        df_tr.drop(columns=["EOD"], inplace=True)
        df_tr = df_tr.sort_values(["date", "symbol", "action"]).reset_index(drop=True)

    final_val = float(df_eq["value"].iloc[-1]) if not df_eq.empty else float(initial_capital)
    years = max(1e-9, (end_date - start_date).days / 365)
    total_ret = (final_val / initial_capital - 1) * 100
    cagr = ((final_val / initial_capital) ** (1 / years) - 1) * 100
    maxdd = (
        ((df_eq["value"].cummax() - df_eq["value"]) / df_eq["value"].cummax()).max() * 100
        if not df_eq.empty else 0.0
    )

    metrics = {
        "invested_capital": float(initial_capital),
        "final_capital": float(round(final_val, 2)),
        "total_return": float(round(total_ret, 2)),
        "gagr": float(round(cagr, 2)),
        "max_drawdown": float(round(maxdd, 2)),
        "filter_action": FILTER_ACTION,
    }

    # ------------------------------------------------------
    # FIFO Closed Trades Analysis
    # ------------------------------------------------------
    # ------------------------------------------------------
    # FIFO Closed Trades Analysis (with intra-holding max/min tracking)
    # ------------------------------------------------------
    def _closed_trades_fifo(tr: pd.DataFrame, exclude_symbols=("gold_bees_1", "gold_bees")):
        """
        Build FIFO closed-trades list. Now supports 'gold_bees' by falling back to df_gold
        when per-symbol history isn't available in stock_cache.
        """
        if tr.empty:
            return pd.DataFrame(columns=[
                "symbol", "buy_date", "sell_date", "buy_price", "sell_price",
                "qty", "roi", "holding_days",
                "peak_profit_percent", "peak_drawdown_percent",
                "peak_price", "bottom_price"
            ])

        df = tr.copy()
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df["shares"] = pd.to_numeric(df["shares"], errors="coerce").fillna(0).astype(int)
        df["price"] = pd.to_numeric(df["price"], errors="coerce")

        if exclude_symbols:
            df = df[~df["symbol"].isin(exclude_symbols)]

        df = df[df["action"].str.startswith(("BUY", "SELL"), na=False)].sort_values(["symbol", "date"]).reset_index(drop=True)

        closed = []
        for sym, g in df.groupby("symbol", sort=False):
            fifo = []

            # primary lookup - per-stock history in stock_cache
            hist = stock_cache.get(sym)

            # fallback: if this is gold_bees, use df_gold (index-style)
            if (hist is None or hist.empty) and ('gold_bees' in str(sym).lower()):
                # create a small DataFrame consistent with stock history expected columns
                if df_gold is not None and not df_gold.empty:
                    hist = df_gold.copy()
                    # unify column name to ch_closing_price expected by the rest of the logic
                    if "ch_closing_price" not in hist.columns and "i_close_price" in hist.columns:
                        hist = hist.rename(columns={"i_close_price": "ch_closing_price"})
                    # some df_gold may have other naming — try to coerce
                    if "ch_closing_price" not in hist.columns:
                        # try any numeric column as last resort
                        numcols = hist.select_dtypes(include=[np.number]).columns.tolist()
                        if numcols:
                            hist["ch_closing_price"] = hist[numcols[0]]
                    # ensure index is datetime
                    hist.index = pd.to_datetime(hist.index)
                else:
                    hist = pd.DataFrame()

            if hist is None or hist.empty:
                # if no history, we can still generate closed trades with NaN peaks
                hist = pd.DataFrame()

            # Ensure closing price column exists before using .astype(float)
            has_close = "ch_closing_price" in hist.columns

            for _, row in g.iterrows():
                act, qty, px, dt = row["action"], int(row["shares"]), float(row["price"]), row["date"]

                if act.startswith("BUY"):
                    fifo.append({"date": dt, "price": px, "qty": qty})

                elif act.startswith("SELL"):
                    sell_qty = qty
                    while sell_qty > 0 and fifo:
                        lot = fifo[0]
                        use = min(lot["qty"], sell_qty)
                        buy_date = lot["date"]
                        sell_date = dt
                        buy_price = lot["price"]
                        sell_price = px

                        # 🔍 Slice historical prices between buy and sell dates if history present
                        try:
                            if not hist.empty and has_close:
                                period_data = hist.loc[(hist.index >= buy_date) & (hist.index <= sell_date)]
                                prices = period_data["ch_closing_price"].astype(float)
                                if not prices.empty:
                                    peak_price = prices.max()
                                    bottom_price = prices.min()
                                    peak_profit = ((peak_price - buy_price) / buy_price) * 100
                                    peak_drawdown = ((bottom_price - buy_price) / buy_price) * 100
                                else:
                                    peak_profit = np.nan
                                    peak_drawdown = np.nan
                                    peak_price = np.nan
                                    bottom_price = np.nan
                            else:
                                peak_profit = np.nan
                                peak_drawdown = np.nan
                                peak_price = np.nan
                                bottom_price = np.nan
                        except Exception:
                            peak_profit = np.nan
                            peak_drawdown = np.nan
                            peak_price = np.nan
                            bottom_price = np.nan

                        # ROI (sell vs buy)
                        try:
                            roi = (sell_price - buy_price) / buy_price * 100.0
                        except Exception:
                            roi = np.nan

                        closed.append({
                            "symbol": sym,
                            "buy_date": buy_date,
                            "sell_date": sell_date,
                            "buy_price": buy_price,
                            "sell_price": sell_price,
                            "qty": use,
                            "roi": roi,
                            "pnl" : (sell_price - buy_price) * use,
                            "holding_days": (sell_date - buy_date).days if pd.notna(buy_date) and pd.notna(sell_date) else None,
                            "peak_profit_percent": peak_profit,
                            "peak_drawdown_percent": peak_drawdown,
                            "peak_price": peak_price,
                            "bottom_price": bottom_price
                        })

                        lot["qty"] -= use
                        sell_qty -= use
                        if lot["qty"] <= 0:
                            fifo.pop(0)

        closed_df = pd.DataFrame(closed)
        if not closed_df.empty:
            closed_df = closed_df.round(3)
        return closed_df


    closed_df = _closed_trades_fifo(df_tr, exclude_symbols=("gold_bees_1",))
    if SAVE_EXCEL and not closed_df.empty:
        today_str = datetime.now().strftime("%Y-%m-%d")
        closed_df.to_csv(f"holding_peaks_{today_str}.csv", index=False)
        print(f"📁 holding_peaks_{today_str}.csv saved")

    # ------------------------------------------------------
    # Extended Performance Metrics
    # ------------------------------------------------------
    if not closed_df.empty:
        wins = closed_df[closed_df["roi"] > 0]
        losses = closed_df[closed_df["roi"] < 0]
        total = len(closed_df)
        win_rate = (len(wins) / total) * 100.0 if total else 0.0
        avg_win = wins["roi"].mean() if not wins.empty else 0.0
        avg_loss = losses["roi"].mean() if not losses.empty else 0.0
        bw = wins.loc[wins["roi"].idxmax()] if not wins.empty else None
        bl = losses.loc[losses["roi"].idxmin()] if not losses.empty else None
        rr = (
            (avg_win * (win_rate / 100.0)) /
            (abs(avg_loss) * (1 - (win_rate / 100.0)))
            if not losses.empty and 0 < win_rate < 100
            else None
        )
        years_bt = max(1e-9, (end_date - start_date).days / 365)
        avg_trades_per_year = total / years_bt

        metrics.update({
            "total_trades": int(total),
            "no_of_winners": int(len(wins)),
            "no_of_losers": int(len(losses)),
            "win_rate_percent": float(round(win_rate, 3)),
            "avg_winners_roi_percent": float(round(avg_win, 3)),
            "avg_losers_roi_percent": float(round(avg_loss, 3)),
            "biggest_winner_roi_percent": float(round(bw["roi"], 3)) if bw is not None else 0.0,
            "biggest_winner_stock": bw["symbol"] if bw is not None else None,
            "biggest_winner_holding_days": int(bw["holding_days"]) if bw is not None else 0,
            "biggest_winner_holding_start_date": bw["buy_date"].strftime("%Y-%m-%d") if bw is not None else None,
            "biggest_winner_holding_end_date": bw["sell_date"].strftime("%Y-%m-%d") if bw is not None else None,
            "biggest_winner_holding_buy_price": float(round(float(bw["buy_price"]), 2)) if bw is not None else None,
            "biggest_winner_holding_sell_price": float(round(float(bw["sell_price"]), 2)) if bw is not None else None,
            "biggest_loser_roi_percent": float(round(bl["roi"], 3)) if bl is not None else 0.0,
            "biggest_loser_stock": bl["symbol"] if bl is not None else None,
            "biggest_loser_holding_days": int(bl["holding_days"]) if bl is not None else 0,
            "biggest_loser_holding_start_date": bl["buy_date"].strftime("%Y-%m-%d") if bl is not None else None,
            "biggest_loser_holding_end_date": bl["sell_date"].strftime("%Y-%m-%d") if bl is not None else None,
            "biggest_loser_holding_buy_price": float(round(float(bl["buy_price"]), 2)) if bl is not None else None,
            "biggest_loser_holding_sell_price": float(round(float(bl["sell_price"]), 2)) if bl is not None else None,
            "risk_reward": float(round(rr, 3)) if rr is not None else None,
            "avg_trades_per_year": float(round(avg_trades_per_year, 2)),
        })
    else:
        metrics.update({
            "total_trades": 0,
            "no_of_winners": 0,
            "no_of_losers": 0,
            "win_rate_percent": 0.0,
            "avg_winners_roi_percent": 0.0,
            "avg_losers_roi_percent": 0.0,
            "biggest_winner_roi_percent": 0.0,
            "biggest_winner_stock": None,
            "biggest_winner_holding_days": 0,
            "biggest_winner_holding_start_date": None,
            "biggest_winner_holding_end_date": None,
            "biggest_winner_holding_buy_price": None,
            "biggest_winner_holding_sell_price": None,
            "biggest_loser_roi_percent": 0.0,
            "biggest_loser_stock": None,
            "biggest_loser_holding_days": 0,
            "biggest_loser_holding_start_date": None,
            "biggest_loser_holding_end_date": None,
            "biggest_loser_holding_buy_price": None,
            "biggest_loser_holding_sell_price": None,
            "risk_reward": None,
            "avg_trades_per_year": 0.0,
        })

    # ------------------------------------------------------
    # Monthly & Equity Curve Stats
    # ------------------------------------------------------
    monthly_json, monthly_stats_json, equity_curve_json, monthly_equity_json, yearly_equity_json = [], [], [], [], []
    try:
        if not closed_df.empty:
            cdf = closed_df.copy()
            cdf["sell_date"] = pd.to_datetime(cdf["sell_date"])

            # Sort trades by close date
            cdf = cdf.sort_values("sell_date")

            # Month Key
            cdf["month"] = cdf["sell_date"].dt.to_period("M").astype(str)

            # Cumulative realized P&L overall
            cdf["cum_pnl"] = cdf["pnl"].cumsum()

            # Total PnL for each month (realized only)
            monthly_pnl = cdf.groupby("month").agg(
                MonthPnL=("pnl", "sum"),
            ).reset_index()

            # Compute start cumulative PnL before each month
            cum_before = []
            for m in monthly_pnl["month"]:
                # last date of previous month
                period = pd.Period(m)
                prev_end = (period - 1).end_time

                # pnl before month start (<= prev_end)
                pnl_before = cdf.loc[cdf["sell_date"] <= prev_end, "pnl"].sum()
                cum_before.append(pnl_before)

            monthly_pnl["CumBeforeMonth"] = cum_before

            initial_capital = initial_capital  #300000  # << change if dynamic

            # Capital base = initial + cumulative profit before this month
            monthly_pnl["Capital_Base"] = initial_capital + monthly_pnl["CumBeforeMonth"]

            # Monthly ROI %
            monthly_pnl["Monthly_ROI_Pct"] = (monthly_pnl["MonthPnL"] / monthly_pnl["Capital_Base"]) * 100

            # Final JSON
            monthly_json = monthly_pnl.round(3).to_dict(orient="records")


        # ======================================================
        # 2) TRUE PORTFOLIO MONTHLY ROI (Replace monthly_json)
        # ======================================================
        if "df_eq_curve" in locals():
            df_e = df_eq_curve.copy()
            df_e.index = pd.to_datetime(df_e.index)

            # Determine first & last value per month
            month_df = df_e.resample("M").agg(
                month_start=("portfolio", lambda x: x.iloc[0]),
                month_end=("portfolio", lambda x: x.iloc[-1])
            )

            # True portfolio ROI
            month_df["Monthly_ROI_Pct"] = ((month_df["month_end"] / month_df["month_start"]) - 1) * 100

            # Build JSON output
            monthly_json = [
                {
                    "month": idx.strftime("%Y-%m"),
                    "Monthly_ROI_Pct": round(row["Monthly_ROI_Pct"], 3),
                    "Start_Value": round(row["month_start"], 2),
                    "End_Value": round(row["month_end"], 2)
                }
                for idx, row in month_df.iterrows()
            ]

        if not df_eq.empty:
            if SAVE_EXCEL and not closed_df.empty:
                today_str = datetime.now().strftime("%Y-%m-%d")
                closed_df.to_csv(f"closed_trades_{today_str}.csv", index=False)
                print(f"📁 closed_trades_{today_str}.csv saved")
            df_eq_curve = df_eq.copy()
            df_eq_curve["date"] = pd.to_datetime(df_eq_curve["date"])
            df_eq_curve["portfolio"] = df_eq_curve["value"].astype(float)

            # ✅ Drop duplicate dates before reindexing (keep last record of each date)
            df_eq_curve = df_eq_curve.drop_duplicates(subset=["date"], keep="last")

            full_dates = pd.to_datetime(df_index.index.unique())
            full_dates = full_dates[(full_dates >= start_date) & (full_dates <= end_date)]

            # ✅ Reindex safely
            df_eq_curve = df_eq_curve.set_index("date").reindex(full_dates, method="ffill").sort_index()

            df_eq_curve["portfolio"] = df_eq_curve["portfolio"].ffill().bfill()
            df_eq_curve["date"] = df_eq_curve.index
            
            # ======================================================
            # 🎯 Recalculate CAGR, Max Drawdown, and Detailed Drawdown Stats
            # ======================================================
            try:
                equity_series = df_eq_curve["portfolio"].astype(float).copy()
                equity_series.index = pd.to_datetime(df_eq_curve.index)

                # ✅ CAGR (annualized)
                initial_val = equity_series.iloc[0]
                final_val = equity_series.iloc[-1]
                years = (equity_series.index[-1] - equity_series.index[0]).days / 365.25
                cagr = ((final_val / initial_val) ** (1 / years) - 1) * 100

                # ✅ Max Drawdown (Daily Accurate)
                rolling_max = equity_series.cummax()
                drawdown = (equity_series - rolling_max) / rolling_max * 100
                maxdd = drawdown.min()  # negative (e.g., -9.87)
                maxdd_abs = abs(maxdd)

                # ✅ Identify Drawdown Start (Peak) and End (Bottom)
                dd_end = drawdown.idxmin()  # date of lowest drawdown
                dd_start_candidates = equity_series.loc[:dd_end][equity_series.loc[:dd_end] == rolling_max.loc[:dd_end]]
                dd_start = dd_start_candidates.index[-1] if not dd_start_candidates.empty else equity_series.index[0]

                # ✅ Portfolio values at start and bottom
                start_value = equity_series.loc[dd_start]
                bottom_value = equity_series.loc[dd_end]
                drop_amount = start_value - bottom_value
                drop_percent = (drop_amount / start_value) * 100

                # ✅ Drawdown recovery (optional)
                after_dd = equity_series.loc[dd_end:]
                dd_recover = after_dd[after_dd >= start_value].index
                dd_recover_date = dd_recover[0] if len(dd_recover) > 0 else None

                # ✅ Duration in days
                dd_duration_days = (
                    (dd_recover_date - dd_start).days if dd_recover_date is not None else (equity_series.index[-1] - dd_start).days
                )


                # ======================================================
                # 📌 DAILY DRAWDOWN JSON (FOR CHART DISPLAY)
                # ======================================================
                # Build dd_df first (shared by weekly/monthly blocks below)
                dd_df = pd.DataFrame({
                    "date": pd.to_datetime(equity_series.index),
                    "portfolio_value": equity_series.values,
                    "peak_value": rolling_max.values
                })
                dd_df["drawdown_rupee"] = dd_df["portfolio_value"] - dd_df["peak_value"]
                dd_df["drawdown_pct"] = (dd_df["drawdown_rupee"] / dd_df["peak_value"]) * 100

                try:
                    # date_str column only for JSON output — dd_df keeps datetime for weekly/monthly
                    dd_export = dd_df.copy()
                    dd_export["date"] = dd_export["date"].dt.strftime("%Y-%m-%d")
                    daily_drawdown_json = dd_export.round(4).to_dict(orient="records")
                except Exception as e:
                    print(f"[Error: Daily Drawdown JSON] {e}")
                    daily_drawdown_json = []
                
                
                # ======================================================
                # 📌 WEEKLY DRAWDOWN SUMMARY (Compressed from Daily)
                # ======================================================
                try:
                    dd_daily = dd_df.copy()  # date column is already datetime
                    dd_daily["week_start"] = dd_daily["date"] - pd.to_timedelta(dd_daily["date"].dt.weekday, unit='D')
                    dd_daily["week_end"] = dd_daily["week_start"] + pd.Timedelta(days=4)

                    weekly_dd = []

                    for week, week_data in dd_daily.groupby("week_start"):
                        week_data = week_data.sort_values("date")

                        # lowest point in week (maximum drawdown inside that week)
                        bottom_row = week_data.loc[week_data["drawdown_pct"].idxmin()]

                        # peak up to the bottom date (this is global peak tracking)
                        peak_value = bottom_row["peak_value"]
                        bottom_value = bottom_row["portfolio_value"]

                        drawdown_rupee = bottom_value - peak_value
                        drawdown_pct = (drawdown_rupee / peak_value) * 100

                        weekly_dd.append({
                            "week": f"{week.strftime('%Y-%m-%d')}/{(week + pd.Timedelta(days=4)).strftime('%Y-%m-%d')}",
                            "start_date": week.strftime("%Y-%m-%d"),
                            "end_date": (week + pd.Timedelta(days=4)).strftime("%Y-%m-%d"),
                            "peak_value": round(peak_value, 2),
                            "bottom_value": round(bottom_value, 2),
                            "drawdown_rupee": round(drawdown_rupee, 2),
                            "drawdown_pct": round(drawdown_pct, 4),
                            "date" : (week + pd.Timedelta(days=4)).strftime("%Y-%m-%d"),
                        })

                    weekly_drawdown_json = weekly_dd
                except Exception as e:
                    print(f"[Error: Weekly Drawdown JSON] {e}")
                    weekly_drawdown_json = []
                    
                # ======================================================
                # 📌 MONTHLY DRAWDOWN SUMMARY (Compressed from Daily)
                # ======================================================

                try:
                    dd_month = dd_df.copy()  # date column is already datetime
                    dd_month["month"] = dd_month["date"].dt.to_period("M")

                    monthly_dd = []

                    for month, mdf in dd_month.groupby("month"):
                        mdf = mdf.sort_values("date")

                        # Use month-END drawdown so recovery to all-time high shows as 0%
                        end_row = mdf.iloc[-1]
                        peak_value = end_row["peak_value"]
                        end_value = end_row["portfolio_value"]
                        drawdown_pct = end_row["drawdown_pct"]  # (end - peak) / peak * 100

                        # Worst intra-month drawdown (for reference only)
                        worst_row = mdf.loc[mdf["drawdown_pct"].idxmin()]

                        monthly_dd.append({
                            "month": str(month),
                            "start_date": mdf["date"].min().strftime("%Y-%m-%d"),
                            "end_date": mdf["date"].max().strftime("%Y-%m-%d"),
                            "peak_value": round(peak_value, 2),
                            "bottom_value": round(end_value, 2),
                            "drawdown_rupee": round(end_value - peak_value, 2),
                            "drawdown_pct": round(drawdown_pct, 4),
                            "worst_intra_month_pct": round(float(worst_row["drawdown_pct"]), 4),
                            "date": mdf["date"].max().strftime("%Y-%m-%d"),
                        })

                    monthly_drawdown_json = monthly_dd

                except Exception as e:
                    print(f"[Error: Monthly Drawdown JSON] {e}")
                    monthly_drawdown_json = []

                # ✅ Print nicely
                print("\n📊 DETAILED DRAWDOWN REPORT")
                print(f"🏁 Start (Peak): {dd_start.date()} → Portfolio ₹{start_value:,.2f}")
                print(f"📉 Bottom:       {dd_end.date()} → Portfolio ₹{bottom_value:,.2f}")
                print(f"💸 Drop:         ₹{drop_amount:,.2f} ({drop_percent:.2f}%)")
                print(f"⏳ Duration:     {dd_duration_days} days")
                if dd_recover_date:
                    print(f"🔁 Recovered by: {dd_recover_date.date()}")

                # ✅ Update metrics
                metrics.update({
                    "gagr": float(round(cagr, 2)),
                    "max_drawdown": float(round(maxdd_abs, 2)),
                    "drawdown_start_date": dd_start.strftime("%Y-%m-%d"),
                    "drawdown_bottom_date": dd_end.strftime("%Y-%m-%d"),
                    "drawdown_recovery_date": dd_recover_date.strftime("%Y-%m-%d") if dd_recover_date else None,
                    "drawdown_duration_days": int(dd_duration_days),
                    "drawdown_start_portfolio": float(round(start_value, 2)),
                    "drawdown_bottom_portfolio": float(round(bottom_value, 2)),
                    "drawdown_drop_amount": float(round(drop_amount, 2)),
                    "drawdown_drop_percent": float(round(drop_percent, 2)),
                    "daily_drawdown_json": daily_drawdown_json,
                    "monthly_drawdown_json": monthly_drawdown_json,
                })

            
            except Exception as e:
                print(f"[Warning] Daily metric recompute failed: {e}")

            if "i_close_price" in df_index.columns:
                idx_col = "i_close_price"
            elif "ch_closing_price" in df_index.columns:
                idx_col = "ch_closing_price"
            else:
                idx_col = df_index.select_dtypes(include=[np.number]).columns[0]

            df_idx_curve = df_index.copy()
            df_idx_curve["index"] = df_idx_curve[idx_col].astype(float)

            # ======================================================
            # 📆 Monthly Equity Curve (End-of-Month Portfolio Snapshot)
            # ======================================================
            merged = pd.merge(
                df_eq_curve[["date", "portfolio"]],
                df_idx_curve[["index"]],
                left_on="date",
                right_index=True,
                how="inner"
            ).round(2)

            # ✅ Convert to month-end data (string Month to avoid JSON errors)
            merged["Month"] = merged["date"].dt.to_period("M").astype(str)

            # ✅ Pick last available trading day of each month
            month_end_df = (
                merged.groupby("Month")
                .tail(1)
                .sort_values("date")
                .reset_index(drop=True)
            )
            
            # ======================================================
            # 💰 True Monthly PNL & ROI (based on month start-end values)
            # ======================================================

            # Get the first trading day value for each month
            month_start_df = (
                merged.groupby("Month")
                .head(1)[["Month", "portfolio"]]
                .rename(columns={"portfolio": "start_portfolio"})
            )

            # Merge Start & End values
            month_calc = pd.merge(
                month_end_df,
                month_start_df,
                on="Month",
                how="left"
            )
            # ➕ Copy other computed columns from month_end_df into month_calc
            extra_cols = [
                "month_return_pct",
                "realized_return_pct",
                "unrealized_return_pct",
                "goldbees_realized_pct",
                "goldbees_unrealized_pct"
            ]

            for col in extra_cols:
                if col in month_end_df.columns:
                    month_calc[col] = month_end_df[col].values
                else:
                    month_calc[col] = 0  # default if missing


            # A) Monthly PNL (Absolute ₹)
            month_calc["monthly_pnl"] = month_calc["portfolio"] - month_calc["start_portfolio"]

            # B) Monthly ROI % (within the month)
            month_calc["monthly_roi_pct"] = (
                (month_calc["monthly_pnl"] / month_calc["start_portfolio"]) * 100
            ).round(3)


            # ✅ Format for JSON output
            month_end_df["date"] = month_end_df["date"].dt.strftime("%Y-%m-%d")
            # ======================================================
            # 🔥 Correct Equity-Based Monthly ROI (exact match with total return)
            # ======================================================
            month_end_df["portfolio"] = month_end_df["portfolio"].astype(float)

            # Shift to get previous month equity
            month_end_df["prev_portfolio"] = month_end_df["portfolio"].shift(1)

            # ROI formula
            month_end_df["month_return_pct"] = (
                (month_end_df["portfolio"] / month_end_df["prev_portfolio"] - 1) * 100
            )

            # First month should compare to initial capital, not NaN
            first_idx = month_end_df.index[0]
            month_end_df.loc[first_idx, "month_return_pct"] = (
                (month_end_df.loc[first_idx, "portfolio"] / float(initial_capital) - 1) * 100
            )

            # Remove helper column
            month_end_df.drop(columns=["prev_portfolio"], inplace=True)
            # ======================================================
            # 🎯 NEW — Add Realised / GoldBees / Unrealised Breakdown
            # ======================================================

            # ---- Realised (All closed trades) ----
            if not closed_df.empty:
                closed_df["month"] = pd.to_datetime(closed_df["sell_date"]).dt.to_period("M").astype(str)
                realized_map = closed_df.groupby("month")["roi"].sum().to_dict()
            else:
                realized_map = {}

            month_end_df["realized_return_pct"] = month_end_df["Month"].map(realized_map).fillna(0.0)


            # ---- GoldBees Realised ----
            if not closed_df.empty:
                gb_df = closed_df[closed_df["symbol"].str.lower() == "gold_bees"]
                goldbees_realized_map = gb_df.groupby("month")["roi"].sum().to_dict()
            else:
                goldbees_realized_map = {}

            month_end_df["goldbees_realized_pct"] = month_end_df["Month"].map(goldbees_realized_map).fillna(0.0)


            # ---- GoldBees Unrealised MTM % ----
            if not df_eq_curve.empty and "gold_bees_value" in df_eq_curve.columns:
                gold_df = df_eq_curve.copy()
                gold_df["Month"] = gold_df.index.to_period("M").astype(str)
                gold_month_mtm = gold_df.groupby("Month")["gold_bees_value"].last().to_dict()

                gold_prev_mtm = {}
                keys = list(gold_month_mtm.keys())
                for i in range(len(keys)):
                    prev_key = keys[i - 1] if i > 0 else None
                    gold_prev_mtm[keys[i]] = gold_month_mtm.get(prev_key, 0)
            else:
                gold_month_mtm = {}
                gold_prev_mtm = {}

            def compute_gold_unrealized(m):
                cur_val = gold_month_mtm.get(m, 0)
                prev_val = gold_prev_mtm.get(m, 0)
                delta = cur_val - prev_val
                return round((delta / initial_capital) * 100, 3) if delta != 0 else 0.0

            month_end_df["goldbees_unrealized_pct"] = month_end_df["Month"].apply(compute_gold_unrealized)


            # ---- Non-Gold Unrealised (Everything else) ----
            month_end_df["unrealized_return_pct"] = (
                month_end_df["month_return_pct"]
                - month_end_df["realized_return_pct"]
                - month_end_df["goldbees_realized_pct"]
                - month_end_df["goldbees_unrealized_pct"]
            ).round(3)

            # ======================================================
            # 📌 Prepare BOTH objects: monthly_equity_json + monthly_json (alias)
            # ======================================================
            monthly_equity_json = month_calc[[
                "Month",
                "date",
                "start_portfolio",
                "portfolio",
                "monthly_pnl",
                "monthly_roi_pct",
                "month_return_pct",
                "realized_return_pct",
                "unrealized_return_pct",
                "goldbees_realized_pct",
                "goldbees_unrealized_pct"
            ]].round(3).to_dict(orient="records")
            
            # ======================================================
            # 🟦 YEAR-WISE PNL & ROI (based on first & last day of each year)
            # ======================================================

            # Convert date column to datetime if not already
            month_calc["date"] = pd.to_datetime(month_calc["date"])

            # Extract year
            month_calc["Year"] = month_calc["date"].dt.year.astype(str)

            # 1️⃣ Get the FIRST portfolio value of each year
            year_start_df = (
                month_calc.groupby("Year")
                .head(1)[["Year", "start_portfolio"]]
                .rename(columns={"start_portfolio": "year_start_portfolio"})
            )

            # 2️⃣ Get the LAST portfolio value of each year
            year_end_df = (
                month_calc.groupby("Year")
                .tail(1)[["Year", "portfolio"]]
                .rename(columns={"portfolio": "year_end_portfolio"})
            )

            # 3️⃣ Merge into a consolidated table
            year_df = pd.merge(year_start_df, year_end_df, on="Year", how="inner")

            # 4️⃣ Calculate Yearly PNL
            year_df["yearly_pnl"] = year_df["year_end_portfolio"] - year_df["year_start_portfolio"]

            # 5️⃣ Calculate Yearly ROI %
            year_df["yearly_roi_pct"] = (
                (year_df["yearly_pnl"] / year_df["year_start_portfolio"]) * 100
            ).round(3)

            # 6️⃣ Convert to JSON
            yearly_equity_json = year_df.round(3).to_dict(orient="records")

            print(f"📌 Generated yearly_equity_json → {len(yearly_equity_json)} rows")





            print(f"📌 Generated monthly_equity_json → {len(monthly_equity_json)} rows")
            print(f"📌 monthly_json alias created (equity based)")

            equity_curve_json = month_end_df.to_dict(orient="records")

            print(f"📊 Generated monthly equity_curve_json with {len(equity_curve_json)} records.")



    except Exception as e:
        print(f"[Monthly/Chart Block Error] {e}")
        monthly_json, monthly_stats_json, equity_curve_json, monthly_equity_json, yearly_equity_json = [], [], [], [], []

    # ------------------------------------------------------
    # Snapshot Frame + End-of-Backtest Row
    # ------------------------------------------------------
    snapshots_df = pd.DataFrame([
        {
            "date": s["date"],
            "cash": s["cash"],
            "holdings_count": s["holdings_count"],
            "total_value": s["total_value"],
            "positions": s["positions"],
        } for s in snapshots
    ])
    # ------------------------------------------------------
    #  📌 Unified Daily PnL & Positions Calculation (with buy_date + extra summary cols)
    # ------------------------------------------------------
    try:
        import numpy as np
        import pandas as pd

        # ==================================================
        # 0️⃣ Build a DataFrame from snapshots for portfolio values
        # ==================================================
        snap_rows = []
        for s in snapshots:
            d = pd.to_datetime(s["date"]).normalize()
            cash_snap = float(s["cash"])
            total_val = float(s["total_value"])
            pos_val = total_val - cash_snap
            snap_rows.append({
                "date": d,
                "portfolio_value": round(total_val, 2),
                "positions_value": round(pos_val, 2),
                "cash": round(cash_snap, 2),
            })

        df_snap = pd.DataFrame(snap_rows)
        if not df_snap.empty:
            # If multiple snapshots in one day, keep last one
            df_snap = (
                df_snap.sort_values(["date"])
                    .groupby("date", as_index=False)
                    .last()
            )

        # ==================================================
        # 1️⃣ Build OPEN position rows from snapshots
        # ==================================================
        # ==================================================
        # 1️⃣ Build OPEN position rows with prev_day_close
        # ==================================================
        pos_rows = []

        df_trades_buy = df_tr.copy()
        df_trades_buy["date"] = pd.to_datetime(df_trades_buy["date"])

        # Precompute buy-lookup for O(log n) access inside snapshot loop
        _buy_only = df_trades_buy[df_trades_buy["action"].str.startswith("BUY")].sort_values("date")
        _buy_lookup = {sym: g.reset_index(drop=True) for sym, g in _buy_only.groupby("symbol")}

        # Get previous day last close for each symbol (searchsorted = fast)
        def get_prev_close(symbol, date):
            if symbol.lower() == "gold_bees":
                df = df_gold
                price_col = "i_close_price"
            else:
                df = stock_cache.get(symbol)
                price_col = "ch_closing_price"
            if df is None or df.empty:
                return None
            pos = df.index.searchsorted(date, side="left") - 1
            if pos < 0:
                return None
            return float(df.iloc[pos][price_col])

        for snap in snapshots:
            d = pd.to_datetime(snap["date"]).normalize()

            for p in snap["positions"]:
                sym = p["symbol"]
                qty = int(p["qty"])
                if qty <= 0:
                    continue

                ltp = float(p["price"])          # current price
                prev_close = get_prev_close(sym, d)

                # Fallback if not present (IPO or first day) → treat prev_close = ltp
                if prev_close is None:
                    prev_close = ltp

                # Find the last BUY trade (O(log n) via precomputed lookup)
                sym_buys = _buy_lookup.get(sym)
                if sym_buys is not None and not sym_buys.empty:
                    pos = sym_buys["date"].searchsorted(d, side="right") - 1
                    if pos >= 0:
                        last_buy = sym_buys.iloc[pos]
                        buy_px = float(last_buy["price"])
                        buy_dt = pd.to_datetime(last_buy["date"]).normalize()
                    else:
                        buy_px = ltp
                        buy_dt = d
                else:
                    buy_px = ltp
                    buy_dt = d

                # 💥 NEW PNL based on prev_close
                pnl_rupee = (ltp - prev_close) * qty
                pnl_pct = ((ltp / prev_close - 1) * 100) if prev_close != 0 else 0.0

                pos_rows.append({
                    "date": d,
                    "symbol": sym,
                    "qty": qty,
                    "side": "OPEN",
                    "buy_date": buy_dt,
                    "buy_price": round(buy_px, 4),
                    "prev_day_close": round(prev_close, 4),     # 👈 NEW COLUMN
                    "ltp_price": round(ltp, 4),
                    "position_value": round(ltp * qty, 2),
                    "pnl_rupee": round(pnl_rupee, 2),
                    "pnl_pct": round(pnl_pct, 4),
                    "is_sell": 0,
                })

        df_positions_open = pd.DataFrame(pos_rows)

        # ==================================================
        # 2️⃣ Build SELL rows (Realized PnL from closed_df)
        # ==================================================
        sell_rows = []
        if not closed_df.empty:
            for _, r in closed_df.iterrows():
                qty = int(r["qty"])
                buy_px = float(r["buy_price"])
                sell_date = pd.to_datetime(r["sell_date"]).normalize()
                sell_px = float(r["sell_price"])
                # 🔎 previous close price from df_daily
                prev_close = get_prev_close(r["symbol"], sell_date)
                # print("prev_close",prev_close,"d",d,'r["symbol"]',r["symbol"], "sell_date",sell_date)
                pnl_rupee = (sell_px - buy_px) * qty
                pnl_pct = ((sell_px / prev_close - 1) * 100) if prev_close != 0 else 0.0
        
                sell_rows.append({
                    "date": pd.to_datetime(r["sell_date"]).normalize(),
                    "symbol": r["symbol"],
                    "qty": qty,
                    "side": "OPEN",
                    "buy_date": pd.to_datetime(r["buy_date"]).normalize(),
                    "buy_price": round(buy_px, 4),
                    "prev_day_close": round(prev_close, 4),     # 👈 NEW COLUMN
                    "ltp_price": round(sell_px, 4),   # LTP = sell price on sell date
                    "position_value": round(sell_px * qty, 2),
                    "pnl_rupee": round((sell_px - prev_close) * qty, 2),
                    "pnl_pct": round(float(pnl_pct), 4),
                    "is_sell": 0,
                })
                sell_rows.append({
                    "date": pd.to_datetime(r["sell_date"]).normalize(),
                    "symbol": r["symbol"],
                    "qty": qty,
                    "side": "SELL",
                    "buy_date": pd.to_datetime(r["buy_date"]).normalize(),
                    "buy_price": round(buy_px, 4),
                    "prev_day_close": round(prev_close, 4),     # 👈 NEW COLUMN
                    "ltp_price": round(sell_px, 4),   # LTP = sell price on sell date
                    "position_value": round(sell_px * qty, 2),
                    "pnl_rupee": round(pnl_rupee, 2),
                    "pnl_pct": round(float(r["roi"]), 4),
                    "is_sell": 1,
                })

        df_positions_sell = pd.DataFrame(sell_rows)


        # ==================================================
        # 🟦 MONTHLY CLOSED TRADE PNL (is_sell = 1)
        # ==================================================
        
            
            
        # ==================================================
        # 3️⃣ Combine OPEN + SELL position details
        # ==================================================
        if not df_positions_open.empty or not df_positions_sell.empty:
            df_positions_detail = pd.concat(
                [df_positions_open, df_positions_sell],
                ignore_index=True
            )
            # reorder column to move prev_day_close next to buy_price
            cols = list(df_positions_detail.columns)
            cols.remove("prev_day_close")
            cols.insert(cols.index("buy_price") + 1, "prev_day_close")
            df_positions_detail = df_positions_detail[cols]

            df_positions_detail = df_positions_detail.sort_values(
                ["date", "side", "symbol", "buy_date"]
            )
        else:
            df_positions_detail = pd.DataFrame()

        # # ==================================================
        # # 4️⃣ Daily PnL from PositionsDetail (matches exactly)
        # # ==================================================
        if not df_positions_detail.empty:
            df_daily = (
                df_positions_detail.groupby("date")["pnl_rupee"]
                                .sum()
                                .reset_index()
            )
        else:
            df_daily = pd.DataFrame(columns=["date", "pnl_rupee"])

        df_daily = df_daily.sort_values("date")
        
        # ==================================================
        # 4️⃣ Daily PnL from PositionsDetail (MTM for Open Positions Only)
        # ==================================================
        if not df_positions_detail.empty:
            # Filter only open positions (is_sell == 0)
            df_open = df_positions_detail[df_positions_detail["is_sell"] == 0]

            df_daily = (
                df_open.groupby("date")["pnl_rupee"]
                    .sum()
                    .reset_index()
            )
        else:
            df_daily = pd.DataFrame(columns=["date", "pnl_rupee"])
        
        
        
        
        df_daily = df_daily.sort_values("date")

        # Cumulative Rupee PnL and Value
        df_daily["cumulative_pnl_rupee"] = df_daily["pnl_rupee"].cumsum()
        df_daily["value"] = float(initial_capital) + df_daily["cumulative_pnl_rupee"]

        # Day PnL % vs previous day portfolio
        df_daily["day_pnl_pct"] = (
            df_daily["pnl_rupee"] / df_daily["value"].shift(1) * 100
        ).fillna(0.0)

        # Cumulative %
        df_daily["cumulative_pnl_pct"] = (
            df_daily["value"] / float(initial_capital) - 1
        ) * 100.0

        # ==================================================
        # 5️⃣ Merge with snapshot-based portfolio values
        # ==================================================
        if not df_snap.empty:
            df_daily = df_daily.merge(df_snap, on="date", how="left")
        else:
            # If for some reason snapshots missing, backfill zeros
            df_daily["portfolio_value"] = df_daily["value"]
            df_daily["positions_value"] = np.nan
            df_daily["cash"] = np.nan

        # Add invested_capital & remaining_capital
        df_daily["invested_capital"] = float(initial_capital)
        # Here remaining_capital = current cash (free capital)
        df_daily["remaining_capital"] = df_daily["cash"]

        # Final formatting
        df_daily["date"] = df_daily["date"].dt.strftime("%Y-%m-%d")
        # print("df_daily before",df_daily.tail(20))
        if not df_positions_detail.empty:
            df_positions_detail["date"] = df_positions_detail["date"].dt.strftime("%Y-%m-%d")
            df_positions_detail["buy_date"] = df_positions_detail["buy_date"].dt.strftime("%Y-%m-%d")

        # ==================================================
        # 6️⃣ Export to Excel (only when SAVE_EXCEL=True)
        # ==================================================
        if SAVE_EXCEL:
            import random
            out_path = f"daily_pnl_matched_with_details{random.randint(1, 10000)}.xlsx"
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df_daily.to_excel(writer, sheet_name="DailySummary", index=False)
                if not df_positions_detail.empty:
                    df_positions_detail.to_excel(writer, sheet_name="PositionsDetail", index=False)
            print(f"📁 Unified Daily PnL Exported → {out_path}")
        # ==================================================
        # 🟦 MONTHLY CLOSED TRADE PNL (is_sell = 1)
        # ==================================================
        if not df_positions_detail.empty:
            df_closed = df_positions_detail[df_positions_detail["is_sell"] == 1].copy()

            # Ensure datetime
            df_closed["date"] = pd.to_datetime(df_closed["date"])

            # Extract month as YYYY-MM
            df_closed["Month"] = df_closed["date"].dt.to_period("M").astype(str)

            # Group by month (realized PnL only)
            monthly_closed = (
                df_closed.groupby("Month")
                        .agg(
                            closed_pnl_rupee=("pnl_rupee", "sum"),
                            closed_roi_pct=("pnl_pct", "sum")  # sum % for all closed trades
                        )
                        .reset_index()
            )

            # Optional rounding & JSON output
            monthly_closed_json = monthly_closed.round(3).to_dict(orient="records")

        else:
            monthly_closed_json = []

        
        # ================================================
        # 🆕 WEEKLY / MONTHLY / YEARLY PNL (from df_daily)
        # ================================================
        try:
            df_daily_pnl = df_daily.copy()
            df_daily_pnl["date"] = pd.to_datetime(df_daily_pnl["date"])
            df_daily_pnl["pnl_rupee"] = pd.to_numeric(df_daily_pnl["pnl_rupee"], errors="coerce").fillna(0)

            # Vectorised prev_value: shift portfolio_value array, fill row-0 with initial_capital
            _pv_arr = df_daily_pnl["portfolio_value"].values.astype(float)
            _pv_prev = np.empty(len(_pv_arr))
            _pv_prev[0] = float(initial_capital)
            _pv_prev[1:] = _pv_arr[:-1]

            # Map date → previous portfolio_value for O(1) lookup in weekly/monthly/yearly loops
            _date_to_prev = dict(zip(df_daily_pnl["date"], _pv_prev))

            def prev_value(start_date):
                v = _date_to_prev.get(pd.Timestamp(start_date))
                return float(v) if v is not None else float(initial_capital)

            # ------------------------------------------------------------
            # 📅 DAILY JSON (modified as per your formula)
            # ------------------------------------------------------------
            daily_json = []
            for i, r in enumerate(df_daily_pnl.itertuples(index=False)):
                start_val = float(_pv_prev[i])
                daily_json.append({
                    "date": str(r.date.date()),
                    "portfolio_start_value": start_val,
                    "portfolio_end_value": float(r.portfolio_value),
                    "portfolio_return": float(r.pnl_rupee),
                    "pnl_rupee": float(r.pnl_rupee),
                    "portfolio_roi": float(r.day_pnl_pct)
                })

            # ------------------------------------------------------------
            # 📅 WEEKLY (Sunday → Saturday) EXACT MTM logic
            # ------------------------------------------------------------
            df_daily_pnl["week"] = df_daily_pnl["date"].dt.to_period("W-SUN").astype(str)
            weekly_json = []
            for wk, g in df_daily_pnl.groupby("week"):
                start_date = g["date"].min()
                end_date   = g["date"].max()

                start_val = prev_value(start_date)
                end_val   = float(g.iloc[-1]["portfolio_value"])
                pnl       = end_val - start_val  # actual MTM change, not pnl_rupee sum

                weekly_json.append({
                    "week": f"{start_date.date()}/{end_date.date()}",
                    "start_date": str(start_date.date()),
                    "end_date": str(end_date.date()),
                    "portfolio_start_value": start_val,
                    "portfolio_end_value": end_val,
                    "portfolio_return": pnl,
                    "pnl_rupee": pnl,
                    "portfolio_roi": (pnl / start_val) * 100 if start_val else 0
                })

            # ------------------------------------------------------------
            # 📅 MONTHLY
            # ------------------------------------------------------------
            df_daily_pnl["Month"] = df_daily_pnl["date"].dt.to_period("M").astype(str)
            monthly_final_json = []
            for m, g in df_daily_pnl.groupby("Month"):
                start_date = g["date"].min()
                end_date   = g["date"].max()

                start_val = prev_value(start_date)
                end_val   = float(g.iloc[-1]["portfolio_value"])
                pnl       = end_val - start_val  # actual MTM change

                monthly_final_json.append({
                    "month": m,
                    "start_date": str(start_date.date()),
                    "end_date": str(end_date.date()),
                    "portfolio_start_value": start_val,
                    "portfolio_end_value": end_val,
                    "portfolio_return": pnl,
                    "pnl_rupee": pnl,
                    "portfolio_roi": (pnl / start_val) * 100 if start_val else 0
                })

            # ---------------------------------
            # 📅 QUARTERLY (Q1-Q4 MTM PnL)
            # ---------------------------------
            df_daily_pnl["Quarter"] = df_daily_pnl["date"].dt.to_period("Q").astype(str)

            quarterly = []
            for q, block in df_daily_pnl.groupby("Quarter"):
                start_date = block["date"].min()
                end_date   = block["date"].max()

                start_val = prev_value(start_date)
                end_val   = float(block.iloc[-1]["portfolio_value"])
                pnl       = end_val - start_val  # actual MTM change

                quarterly.append({
                    "quarter": q,
                    "start_date": str(start_date.date()),
                    "end_date": str(end_date.date()),
                    "portfolio_start_value": start_val,
                    "portfolio_end_value": end_val,
                    "portfolio_return": pnl,
                    "pnl_rupee": pnl,
                    "portfolio_roi": (pnl / start_val) * 100 if start_val else 0
                })

            quarterly_json = quarterly

            # ------------------------------------------------------------
            # 📅 YEARLY
            # ------------------------------------------------------------
            df_daily_pnl["Year"] = df_daily_pnl["date"].dt.year.astype(str)
            yearly_final_json = []
            for y, g in df_daily_pnl.groupby("Year"):
                start_date = g["date"].min()
                end_date   = g["date"].max()

                start_val = prev_value(start_date)
                end_val   = float(g.iloc[-1]["portfolio_value"])
                pnl       = end_val - start_val  # actual MTM change

                yearly_final_json.append({
                    "year": y,
                    "start_date": str(start_date.date()),
                    "end_date": str(end_date.date()),
                    "portfolio_start_value": start_val,
                    "portfolio_end_value": end_val,
                    "portfolio_return": pnl,
                    "pnl_rupee": pnl,
                    "portfolio_roi": (pnl / start_val) * 100 if start_val else 0
                })

        except Exception as e:
            print("[Weekly/Monthly/Yearly Computation Error]", e)
            daily_json, weekly_json, monthly_final_json, yearly_final_json, quarterly_json = [], [], [], [], []
        
    except Exception as e:
        print(f"[Daily PnL Export Error] {e}")

    
    if not df_tr.empty and not df_eq.empty:
        df_tr.loc[len(df_tr)] = {
            "date": pd.to_datetime(df_eq["date"].iloc[-1]),
            "symbol": "TOTAL",
            "action": "END_OF_BACKTEST",
            "shares": "",
            "price": "",
            "cash_after": df_tr["cash_after"].iloc[-1],
            "cumulative_value": float(round(df_eq["value"].iloc[-1], 2)),
            "remining_capital": df_tr["remining_capital"].iloc[-1],
        }

    print("✅ ultra_backtest_v2_fastest_fixed_v3 completed successfully!")
    print(f"🕒 Elapsed: {time.time() - t0:.2f}s")
    print(f"📊 Regime Events: {len(regime_events)} events → {regime_events}")

    return metrics, df_eq, df_tr, monthly_json, monthly_stats_json, equity_curve_json, snapshots_df, monthly_equity_json, yearly_equity_json, daily_json, weekly_json, monthly_final_json, yearly_final_json, quarterly_json, monthly_closed_json, closed_df, regime_events

